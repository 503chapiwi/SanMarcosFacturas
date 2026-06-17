import pdfplumber
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import streamlit as st
import re
import io

# --- HELPER FUNCTIONS ---
def normalize_text(text):
    if not text: return ""
    nfd = unicodedata.normalize('NFD', str(text))
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn').lower()

def squish_text(text):
    """Aggressively removes ALL spaces, punctuation, hyphens, and hidden characters for a 100% reliable match."""
    if not text: return ""
    t = normalize_text(text)
    return re.sub(r'[^a-z0-9]', '', t)

def safe_float(val):
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-': return 0.0
    s = s.replace(',', '')
    s = re.sub(r'[^\d\.\-]', '', s)
    if s.count('.') > 1:
        parts = s.rsplit('.', 1)
        s = parts[0].replace('.', '') + '.' + parts[1]
    try: return float(s)
    except ValueError: return 0.0

def clean_currency(value):
    if not value: return 0.0
    raw = str(value).strip().replace(' ', '')
    raw = re.sub(r'[^\d\.,]', '', raw)
    if not raw: return 0.0

    if re.search(r',\d{1,2}$', raw):
        parts = raw.rsplit(',', 1)
        raw = parts[0].replace('.', '').replace(',', '') + '.' + parts[1]
    else:
        raw = raw.replace(',', '')

    if raw.count('.') > 1:
        parts = raw.rsplit('.', 1)
        raw = parts[0].replace('.', '') + '.' + parts[1]

    try: return float(raw)
    except ValueError: return 0.0

def get_master_cell(ws, r_idx, c_idx):
    cell = ws.cell(row=r_idx, column=c_idx)
    if type(cell).__name__ == 'MergedCell':
        for m_range in ws.merged_cells.ranges:
            if cell.coordinate in m_range:
                return ws.cell(row=m_range.min_row, column=m_range.min_col)
    return cell

def extract_grand_total(tables, text):
    """
    Return the invoice's grand total from the 'TOTALES' row.

    Strategy: find the row whose first cell starts with 'TOTAL' and take the
    LARGEST positive number in that row. This handles both invoice formats:
    - Standard IVA facturas: Total (Q) > Impuestos (Impuestos is ~10.7% of Total
      under Guatemala's 12% IVA), so max() picks the grand total.
    - Pequeño Contribuyente facturas: no Impuestos column (it's empty), so Total
      is the only positive number in the row and max() returns it directly.

    Also handles PDFs where a thin white line inside the TOTALES row causes
    pdfplumber to split it into two rows. When that happens the first fragment
    starts with 'TOTAL' but may be missing the grand-total cell; subsequent
    fragments have an empty first cell. We merge all consecutive continuation
    rows (empty first cell) into the same candidate set before calling max().

    Falls back to text-regex if no TOTAL row is found in the table.
    """
    # PRIMARY: scan tables for a TOTAL-starting row, return the max positive cell.
    # Use an index loop so we can look ahead at continuation rows.
    best = 0.0
    i = 0
    while i < len(tables):
        row = tables[i]
        i += 1
        if not row:
            continue
        first_cell = str(row[0] or '').strip().upper()
        if not first_cell.startswith('TOTAL'):
            continue
        # Collect cells from this row plus any immediately following rows whose
        # first cell is empty/None (split continuation due to a white ruling line).
        combined = list(row)
        while i < len(tables):
            next_row = tables[i]
            if not next_row:
                i += 1
                continue
            next_first = str(next_row[0] or '').strip()
            if next_first == '':
                combined.extend(next_row)
                i += 1
            else:
                break
        positives = [clean_currency(c) for c in combined if c is not None]
        positives = [n for n in positives if n > 0]
        if positives:
            cand = max(positives)
            if cand > best:
                best = cand
    if best > 0:
        return best

    # FALLBACK: regex on raw text. Any 'TOTALES …' line, take its largest number.
    for m in re.finditer(r'(?im)^\s*TOTALES?:?\s*(.+)$', text):
        nums = re.findall(r'[\d.,]+', m.group(1))
        positives = [clean_currency(n) for n in nums]
        positives = [n for n in positives if n > 0]
        if positives:
            cand = max(positives)
            if cand > best:
                best = cand
    return best

# --- TRUCO CSS PARA TRADUCIR LA INTERFAZ A ESPAÑOL ---
st.markdown("""
    <style>
        div[data-testid="stFileUploader"] label p {
            font-size: 40px !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- WEB UI ---
st.title("🇬🇹 MAGA: Procesador de Facturas por la LAE: San Marcos")

# Municipality selector - user must specify which municipality the receipts belong to
MUNICIPIOS_OPCIONES = {
    "Ayutla": 1,
    "Catarina": 2,
    "Comitancillo": 3,
    "Concepcion Tutuapa": 4,
    "El Quetzal": 5,
    "El Tumbador": 6,
    "Esquipulas Palo Gordo": 7,
    "Ixchiguan": 8,
    "La Blanca": 9,
    "La Reforma": 10,
    "Malacatan": 11,
    "Nuevo Progreso": 12,
    "Ocos": 13,
    "Pajapita": 14,
    "Rio Blanco": 15,
    "San Antonio Sacatepequez": 16,
    "San Cristobal Cucho": 17,
    "San Jose El Rodeo": 18,
    "San Jose Ojetenam": 19,
    "San Lorenzo": 20,
    "San Marcos": 21,
    "San Miguel Ixtahuacan": 22,
    "San Pablo": 23,
    "San Pedro Sacatepequez": 24,
    "San Rafael Pie de la Cuesta": 25,
    "Sibinal": 26,
    "Sipacapa": 27,
    "Tacana": 28,
    "Tajumulco": 29,
    "Tejutla": 30,
}

selected_municipio = st.selectbox(
    label='1. Seleccione el Municipio de las facturas',
    options=["-- Seleccionar municipio --"] + list(MUNICIPIOS_OPCIONES.keys()),
    help="Todas las facturas que suba deben corresponder a este municipio"
)

uploaded_pdfs = st.file_uploader(label='2. Seleccione sus Facturas (PDFs)', type='pdf', accept_multiple_files=True)
uploaded_xlsx = st.file_uploader(label='3. Seleccione su Archivo de Excel', type='xlsx')

municipio_valido = selected_municipio != "-- Seleccionar municipio --"

if municipio_valido:
    st.info(f"📍 Municipio seleccionado: **{selected_municipio}**. Asegúrese de que todas las facturas correspondan a este municipio.")
else:
    st.warning("⚠️ Por favor seleccione un municipio antes de iniciar el proceso.")

if st.button("INICIAR PROCESO") and uploaded_pdfs and uploaded_xlsx and municipio_valido:
    try:
        user_m_id = MUNICIPIOS_OPCIONES[selected_municipio]
        user_m_name = selected_municipio

        input_buffer = io.BytesIO(uploaded_xlsx.read())
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active

        # "Extra Detalles" sheet (no Alerta column — no classification anymore)
        if "Extra Detalles" not in wb.sheetnames:
            ws_det = wb.create_sheet("Extra Detalles")
            ws_det.append(['Archivo PDF', 'Nombre Emisor', 'NIT Emisor', 'NIT Receptor', 'Num. DTE', 'Municipio'])
        else:
            ws_det = wb["Extra Detalles"]

        # Collect DTEs already in "Extra Detalles" (Num. DTE is column 5) so we can skip duplicates.
        existing_dtes = set()
        for row in ws_det.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
            if row[0] is not None:
                existing_dtes.add(str(row[0]).strip())

        # 1. Map Excel Columns dynamically
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=15):
            for cell in row:
                if type(cell).__name__ == 'MergedCell': continue
                if not cell.value: continue
                val = normalize_text(str(cell.value))

                if 'agricultura' in val: col_map['agri'] = cell.column
                if 'escuela' in val or 'establecimiento' in val: col_map['escuelas'] = cell.column
                if 'proveedor' in val or 'productor' in val:
                    base_col, base_row, found_total = cell.column, cell.row, False
                    for r_offset in range(1, 4):
                        for c_offset in range(3):
                            sub_cell = ws.cell(row=base_row + r_offset, column=base_col + c_offset)
                            if sub_cell.value and 'total' in normalize_text(str(sub_cell.value)):
                                col_map['productores'] = sub_cell.column
                                found_total = True
                                break
                        if found_total: break
                    if 'productores' not in col_map: col_map['productores'] = base_col

        if 'agri' not in col_map:
            st.error("No encontré la columna de Agricultura en el Excel.")
            st.stop()

        EXCEL_MAPPINGS = {
    1: "ayutla", 2: "catarina", 3: "comitancillo", 4: "concepcion tutuapa",
    5: "el quetzal", 6: "el tumbador", 7: "esquipulas palo gordo", 8: "ixchiguan",
    9: "la blanca", 10: "la reforma", 11: "malacatan", 12: "nuevo progreso",
    13: "ocos", 14: "pajapita", 15: "rio blanco", 16: "san antonio sacatepequez",
    17: "san cristobal cucho", 18: "san jose el rodeo", 19: "san jose ojetenam", 20: "san lorenzo",
    21: "san marcos", 22: "san miguel ixtahuacan", 23: "san pablo",
    24: "san pedro sacatepequez", 25: "san rafael pie de la cuesta",
    26: "sibinal", 27: "sipacapa", 28: "tacana", 29: "tajumulco",  30: "tejutla"
        }

        # 2. Map Excel Rows to Municipalities
        row_map = {}
        for row_ex in ws.iter_rows(min_row=5, max_row=150):
            row_text = " ".join([str(c.value) for c in row_ex if c.value and type(c).__name__ != 'MergedCell'])
            row_squished = squish_text(row_text)
            for m_id, search_key in EXCEL_MAPPINGS.items():
                if m_id in row_map: continue
                key_squished = squish_text(search_key)
                if key_squished in row_squished:
                    row_map[m_id] = row_ex[0].row

        batch_totals = {m_id: {'total': 0.0, 'emisores': set(), 'receptores': set()} for m_id in MUNICIPIOS_OPCIONES.values()}
        new_count = 0
        skipped_non_standard = []
        skipped_duplicate = []
        skipped_no_total = []
        failed_pdfs = []   # (pdf_name, error_message) for unexpected exceptions
        progress_bar = st.progress(0)

        # 3. Process each PDF
        # CRITICAL: each iteration is isolated in its own try/except so that a single
        # corrupt / unusual PDF can't abort the entire batch and silently drop every
        # receipt that comes after it.
        for i, pdf_file in enumerate(uploaded_pdfs):
            try:
                with pdfplumber.open(pdf_file) as pdf:
                    text = "".join([p.extract_text() or "" for p in pdf.pages])
                    tables = []
                    for p in pdf.pages:
                        t = p.extract_table()
                        if t: tables.extend(t)

                    # VALIDATION: Check if this is a standard SAT factura
                    has_dte = bool(re.search(r'N[úu]mero\s*de\s*DTE', text, re.IGNORECASE))
                    has_autorizacion = bool(re.search(r'N[úu]mero\s*de\s*Autorizaci[óo]n', text, re.IGNORECASE))
                    has_nit_emisor = bool(re.search(r'Nit\s*Emisor', text, re.IGNORECASE))
                    marker_count = sum([has_dte, has_autorizacion, has_nit_emisor])
                    is_standard_factura = marker_count >= 2

                    if not is_standard_factura:
                        skipped_non_standard.append(pdf_file.name)
                        continue

                    dte_m = re.search(r'N[úu]mero\s*de\s*DTE:\s*(\d+)', text, re.IGNORECASE)
                    dte_val = dte_m.group(1) if dte_m else pdf_file.name

                    # Duplicate-DTE guard
                    if str(dte_val).strip() in existing_dtes:
                        skipped_duplicate.append((pdf_file.name, dte_val))
                        st.warning(f"Esta factura (Num. DTE: {dte_val}) no ha sido agregado al archivo de Excel: ya ha sido procesado")
                        continue

                    # Grab the single grand total — no line-item parsing, no product matching.
                    grand_total = extract_grand_total(tables, text)
                    if grand_total <= 0:
                        skipped_no_total.append(pdf_file.name)
                        st.warning(f"No se pudo determinar el total de la factura: {pdf_file.name}")
                        continue

                    # Emisor / receptor metadata for "Extra Detalles"
                    nit_e_match = re.search(r'Emisor:\s*([0-9Kk\-]+)', text, re.I)
                    nit_r_match = re.search(r'Receptor:\s*([0-9Kk\-]+)', text, re.I)
                    name_e_match = re.search(r'(?:Factura(?:\s*Pequeño\s*Contribuyente)?)\s*\n+(.*?)\n+Nit\s*Emisor', text, re.IGNORECASE | re.DOTALL)

                    nit_e = nit_e_match.group(1).strip() if nit_e_match else "N/A"
                    nit_r = nit_r_match.group(1).strip() if nit_r_match else "N/A"
                    raw_name = re.sub(r'\s+', ' ', name_e_match.group(1).strip() if name_e_match else "N/A")
                    name_e = re.split(r'(?i)n[úu]mero\s*de\s*autorizaci[óo]n', raw_name)[0]
                    name_e = re.split(r'(?i)\bserie\b', name_e)[0].strip()

                    m_id = user_m_id
                    m_name = user_m_name

                    batch_totals[m_id]['total'] += grand_total
                    if nit_e != "N/A": batch_totals[m_id]['emisores'].add(nit_e)
                    if nit_r != "N/A": batch_totals[m_id]['receptores'].add(nit_r)

                    ws_det.append([pdf_file.name, name_e, nit_e, nit_r, dte_val, m_name])
                    existing_dtes.add(str(dte_val).strip())
                    new_count += 1
            except Exception as e:
                # One PDF blew up — record it and keep going. Do NOT abort the batch.
                failed_pdfs.append((pdf_file.name, f"{type(e).__name__}: {e}"))

            progress_bar.progress((i + 1) / len(uploaded_pdfs))

        # 4. Write to Main Sheet — grand total goes into the Agricultura column
        for target_m_id, r_idx in row_map.items():
            data = batch_totals.get(target_m_id)
            if not data: continue

            if 'agri' in col_map and data['total'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['agri'])
                target_cell.value = safe_float(target_cell.value) + data['total']

            if 'escuelas' in col_map and len(data['receptores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['escuelas'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['receptores'])

            if 'productores' in col_map and len(data['emisores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['productores'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['emisores'])

        # 5. Format "Extra Detalles"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws_det.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.border = thin_border
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_det.column_dimensions[col_letter].width = max_length + 2

        # 6. Final Export
        output = io.BytesIO()
        wb.save(output)

        # Full accounting so every uploaded PDF is accounted for somewhere
        total_uploaded = len(uploaded_pdfs)
        accounted = new_count + len(skipped_non_standard) + len(skipped_duplicate) + len(skipped_no_total) + len(failed_pdfs)

        st.success(
            f"¡Proceso completado! {new_count} de {total_uploaded} facturas agregadas al Excel.\n\n"
            f"Resumen: {new_count} agregadas · "
            f"{len(skipped_duplicate)} duplicadas · "
            f"{len(skipped_non_standard)} no estándar · "
            f"{len(skipped_no_total)} sin total detectable · "
            f"{len(failed_pdfs)} con error"
        )

        if accounted != total_uploaded:
            st.error(f"⚠️ Discrepancia: {total_uploaded} facturas subidas pero solo {accounted} contabilizadas. Por favor reporte este caso.")

        if skipped_non_standard:
            warning_msg = f"⚠️ **{len(skipped_non_standard)} factura(s) no estándar fueron ignoradas** (proformas, cotizaciones, u otros formatos no oficiales). Estas deben procesarse manualmente:\n\n"
            for pdf_name in skipped_non_standard:
                warning_msg += f"- {pdf_name}\n"
            st.warning(warning_msg)

        if skipped_duplicate:
            dup_msg = f"ℹ️ **{len(skipped_duplicate)} factura(s) duplicada(s) ignoradas** (su Num. DTE ya existía en 'Extra Detalles'):\n\n"
            for pdf_name, dte in skipped_duplicate:
                dup_msg += f"- {pdf_name} (DTE: {dte})\n"
            st.info(dup_msg)

        if skipped_no_total:
            nt_msg = f"⚠️ **{len(skipped_no_total)} factura(s) sin total detectable** — no se encontró fila 'TOTALES'. Procesar manualmente:\n\n"
            for pdf_name in skipped_no_total:
                nt_msg += f"- {pdf_name}\n"
            st.warning(nt_msg)

        if failed_pdfs:
            err_msg = f"❌ **{len(failed_pdfs)} factura(s) fallaron con error inesperado** (revisar manualmente):\n\n"
            for pdf_name, err in failed_pdfs:
                err_msg += f"- {pdf_name} — {err}\n"
            st.error(err_msg)

        output.seek(0)
        st.download_button("Descargar Reporte Final", data=output.getvalue(),
                           file_name="Reporte_MAGA_Actualizado.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Error crítico detectado: {e}")
