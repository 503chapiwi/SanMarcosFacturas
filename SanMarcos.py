import pdfplumber
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import streamlit as st
import re
import io
from rapidfuzz import fuzz, process

# --- HELPER FUNCTIONS ---
def normalize_text(text):
    if not text: return ""
    nfd = unicodedata.normalize('NFD', str(text))
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn').lower()

def squish_text(text):
    """Aggressively removes ALL spaces, punctuation, hyphens, and hidden characters for a 100% reliable match."""
    if not text: return ""
    t = normalize_text(text)
    return re.sub(r'[^a-z0-9]', '', t)

def safe_float(val):
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-': return 0.0
    s = s.replace(',', '') 
    s = re.sub(r'[^\d\.\-]', '', s) 
    if s.count('.') > 1:
        parts = s.rsplit('.', 1)
        s = parts[0].replace('.', '') + '.' + parts[1]
    try: return float(s)
    except ValueError: return 0.0

def clean_currency(value):
    if not value: return 0.0
    raw = str(value).strip().replace(' ', '')
    raw = re.sub(r'[^\d\.,]', '', raw)
    if not raw: return 0.0
    
    if re.search(r',\d{1,2}$', raw):
        parts = raw.rsplit(',', 1)
        raw = parts[0].replace('.', '').replace(',', '') + '.' + parts[1]
    else:
        raw = raw.replace(',', '')
        
    if raw.count('.') > 1:
        parts = raw.rsplit('.', 1)
        raw = parts[0].replace('.', '') + '.' + parts[1]
        
    try: return float(raw)
    except ValueError: return 0.0

def extract_value_from_row(row_list, total_idx):
    if total_idx != -1 and len(row_list) > total_idx:
        val = clean_currency(row_list[total_idx])
        if val > 0: return val
    for item in reversed(row_list):
        val = clean_currency(item)
        if val > 0: return val
    return 0.0

def extract_school_name(text):
    """
    Extracts the school name from the 'Nombre Receptor:' field of a SAT FEL factura.

    In the extracted PDF text the left column (receptor data) and the right column
    (fechas / moneda) share the same lines, e.g.:

        Nombre Receptor: CONSEJO EDUCATIVO, EORM J.M. CANTON Fecha y hora de certificación: ...
        CHUIXCHIMAL
        Moneda: GTQ

    So the name starts after 'Nombre Receptor:' and may continue on following lines
    until a new labeled field appears. Right-column text ('Fecha y hora', 'Moneda:')
    and the alternate-layout 'Dirección comprador:' field are stripped from each line.
    """
    lines = text.split('\n')
    name_parts = []
    started = False
    for line in lines:
        if not started:
            m = re.match(r'\s*Nombre\s*Receptor:\s*(.*)', line, re.IGNORECASE)
            if not m:
                continue
            started = True
            part = m.group(1)
        else:
            # A continuation line; stop at the next labeled field or the items table.
            if line.strip() == '.' or re.match(
                r'(Moneda|#No|NIT|Nit|N[úu]mero|Serie|Fecha|Direcci[óo]n)\b',
                line.strip(), re.IGNORECASE
            ):
                break
            part = line
        part = re.split(r'(?i)Fecha\s*y\s*hora', part)[0]
        part = re.split(r'(?i)Moneda\s*:', part)[0]
        part = re.split(r'(?i)Direcci[óo]n\s*comprador', part)[0]
        part = part.strip()
        if part:
            name_parts.append(part)
        else:
            break
    name = re.sub(r'\s+', ' ', ' '.join(name_parts)).strip()
    name = name.strip('"').strip()  # some names come fully quoted: "ORGANIZACION..."
    return name or "N/A"

def get_master_cell(ws, r_idx, c_idx):
    cell = ws.cell(row=r_idx, column=c_idx)
    if type(cell).__name__ == 'MergedCell':
        for m_range in ws.merged_cells.ranges:
            if cell.coordinate in m_range:
                return ws.cell(row=m_range.min_row, column=m_range.min_col)
    return cell

def fuzzy_match_category(description, cultivados, abarrotes, threshold=80):
    """
    Uses fuzzy matching to categorize a product description.
    Returns: ('agricultura', best_match_word) or ('abarrotes', best_match_word) or ('unmatched', None)
    """
    if not description:
        return ('unmatched', None)
    
    # Normalize and extract words from description
    desc_normalized = normalize_text(description)
    words = desc_normalized.split()
    
    # Try exact matches first (original logic)
    for word in words:
        if word in cultivados:
            return ('agricultura', word)
        if word in abarrotes:
            return ('abarrotes', word)
    
    # If no exact match, try fuzzy matching
    best_agri_match = None
    best_agri_score = 0
    
    for word in words:
        # Skip very short words (less than 3 chars) for fuzzy matching
        if len(word) < 3:
            continue
            
        # Check against cultivados
        match_result = process.extractOne(word, cultivados, scorer=fuzz.ratio)
        if match_result and match_result[1] >= threshold:
            if match_result[1] > best_agri_score:
                best_agri_score = match_result[1]
                best_agri_match = match_result[0]
    
    best_abar_match = None
    best_abar_score = 0
    
    for word in words:
        if len(word) < 3:
            continue
            
        # Check against abarrotes
        match_result = process.extractOne(word, abarrotes, scorer=fuzz.ratio)
        if match_result and match_result[1] >= threshold:
            if match_result[1] > best_abar_score:
                best_abar_score = match_result[1]
                best_abar_match = match_result[0]
    
    # Return the category with the best match
    if best_agri_score > best_abar_score and best_agri_match:
        return ('agricultura', best_agri_match)
    elif best_abar_match:
        return ('abarrotes', best_abar_match)
    else:
        return ('unmatched', None)

# --- TRUCO CSS PARA TRADUCIR LA INTERFAZ A ESPAÑOL ---
st.markdown("""
    <style> 
        div[data-testid="stFileUploader"] label p {
            font-size: 40px !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- WEB UI ---
st.title("🇬🇹 MAGA: Procesador de Facturas por la LAE: San Marcos")
uploaded_pdfs = st.file_uploader(label='1. Seleccione sus Facturas (PDFs)', type='pdf', accept_multiple_files=True)
uploaded_xlsx = st.file_uploader(label='2. Seleccione su Archivo de Excel', type='xlsx')

if st.button("INICIAR PROCESO") and uploaded_pdfs and uploaded_xlsx:
    try:
        input_buffer = io.BytesIO(uploaded_xlsx.read())
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active 
        
        if "Extra Detalles" not in wb.sheetnames:
            ws_det = wb.create_sheet("Extra Detalles")
            ws_det.append(['Nombre Emisor', 'NIT Emisor', 'NIT Receptor', 'Nombre Escuela', 'Num. DTE', 'Municipio', 'Alerta % Abarrotes'])
        else:
            ws_det = wb["Extra Detalles"]
            # Excel files from older runs lack the 'Nombre Escuela' column:
            # insert it after 'NIT Receptor' so old and new rows stay aligned.
            if normalize_text(str(ws_det.cell(row=1, column=4).value or "")) != normalize_text("Nombre Escuela"):
                ws_det.insert_cols(4)
                ws_det.cell(row=1, column=4).value = "Nombre Escuela"
        
        # Create sheet for unmatched items
        if "Items Sin Clasificar" not in wb.sheetnames:
            ws_unmatched = wb.create_sheet("Items Sin Clasificar")
            ws_unmatched.append(['Descripción', 'Municipio', 'Total (Q)', 'Num. DTE'])
        else:
            ws_unmatched = wb["Items Sin Clasificar"]

        # 1. Map Excel Columns dynamically
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=15): 
            for cell in row:
                if type(cell).__name__ == 'MergedCell': continue
                if not cell.value: continue
                val = normalize_text(str(cell.value))
                
                if 'abarrotes' in val: col_map['abar'] = cell.column
                if 'agricultura' in val: col_map['agri'] = cell.column
                if 'escuela' in val or 'establecimiento' in val: col_map['escuelas'] = cell.column
                if 'proveedor' in val or 'productor' in val:
                    base_col, base_row, found_total = cell.column, cell.row, False
                    for r_offset in range(1, 4):
                        for c_offset in range(3):
                            sub_cell = ws.cell(row=base_row + r_offset, column=base_col + c_offset)
                            if sub_cell.value and 'total' in normalize_text(str(sub_cell.value)):
                                col_map['productores'] = sub_cell.column
                                found_total = True
                                break
                        if found_total: break
                    if 'productores' not in col_map: col_map['productores'] = base_col

        if 'abar' not in col_map or 'agri' not in col_map:
            st.error(f"No encontré las columnas base en el Excel.")
            st.stop()

        department_name = 'san marcos'
        # 2. MASTER MUNICIPALITY DICTIONARY
        MUNICIPIOS = {
            1: {"nombre_oficial": "Ayutla"},
            2: {"nombre_oficial": "Catarina"},
            3: {"nombre_oficial": "Comitancillo"},
            4: {"nombre_oficial": "Concepción Tutapa"},
            5: {"nombre_oficial": "El Quetzal"},
            6: {"nombre_oficial": "El Tumbador"},
            7: {"nombre_oficial": "Esquipulas Palo Gordo", "alias_pdf": ["esquipulas palo gordo"]},
            8: {"nombre_oficial": "Ixchiguan"},
            9: {"nombre_oficial": "La Blanca"},
            10: {"nombre_oficial": "La Reforma"},
            11: {"nombre_oficial": "Malacatan"},
            12: {"nombre_oficial": "Nuevo Progreso"},
            13: {"nombre_oficial": "Ocos"},
            14: {"nombre_oficial": "Pajapita"},
            15: {"nombre_oficial": "Rio Blanco"},
            16: {"nombre_oficial": "San Antonio Sacatapequez", "alias_pdf": ["san antonio sacatepequez"]},
            17: {"nombre_oficial": "San Cristobal Cuhco", "alias_pdf": ["san cristobal cucho"]},
            18: {"nombre_oficial": "San Jose El Rodeo", "alias_pdf": [""]},
            19: {"nombre_oficial": "San Jose Ojetenam", "alias_pdf": ["san jose ojetenam"]},
            20: {"nombre_oficial": "San Lorenzo"},
            21: {"nombre_oficial": "San Marcos", "alias_pdf": ["San Marcos", "san marcos san marcos", "san marcos, san marcos"]},
            22: {"nombre_oficial": "San Miguel Ixtahuacan", "alias_pdf": ["san miguel ixtahuacan"]},
            23: {"nombre_oficial": "San Pablo"},
            24: {"nombre_oficial": "San Pedro Sacatapequez", "alias_pdf": ["San Pedro Sacatepequez"]},
            25: {"nombre_oficial": "San Rafael Pie De La Cuesta", "alias_pdf": ["san rafael", "san pie de la cuesta"]},
            26: {"nombre_oficial": "Sibinal"},
            27: {"nombre_oficial": "Sipacapa"},
            28: {"nombre_oficial": "Tacana"},
            29: {"nombre_oficial": "Tajamulco"},
            30: {"nombre_oficial": "Tejutla"}
        }
        
        search_list = []
        for m_id, data in MUNICIPIOS.items():
            for alias in data["alias_pdf"]:
                search_list.append((alias, m_id, data["nombre_oficial"]))
                
        # CORE FIX: Sorts the list so Totonicapán (ID 1) is ALWAYS evaluated last.
        # Within the other municipalities, sorts by length to catch specific names first.
                search_list.sort(key=lambda x: (
            squish_text(x[2]) == squish_text(department_name),
            -len(x[0])
        ))
        
        EXCEL_MAPPINGS = {
            1: "ayutla", 2: "catarina", 3: "camitancillo", 4: "concepcion tutuapa",
            5: "el quetzal", 6: "el tumbador", 7: "esquipulas palo gordo", 8: "ixchiguan", 
            9: "la blanca", 10: "la reforma", 11: "malacatan", 12: "nuevo progreso", 
            13: "ocos", 14: "pajapita", 15: "rio blanco", 16: "san antonio sacatepequez", 
            17: "san cristobal cucho", 18: "san jose el rodeo", 19: "san jose ojetenam", 
            20: "san lorenzo", 21: "san marcos", 22: "san miguel ixtahuacan", 23: "san pablo", 
            24: "san pedro sacatepequez", 25: "san rafel pie de la cuesta", 26: "sibinal", 
            27: "sipacapa", 28: "tacana", 29: "tajamulco", 30: "tejutla", 
        }

        # 3. Map Excel Rows to Municipalities
        row_map = {}
        for row_ex in ws.iter_rows(min_row=5, max_row=150):
            row_text = " ".join([str(c.value) for c in row_ex if c.value and type(c).__name__ != 'MergedCell'])
            row_squished = squish_text(row_text)
            for m_id, search_key in EXCEL_MAPPINGS.items():
                if m_id in row_map: continue
                key_squished = squish_text(search_key)
                if key_squished in row_squished:
                    row_map[m_id] = row_ex[0].row

        batch_totals = {m_id: {'abar': 0.0, 'agri': 0.0, 'emisores': set(), 'receptores': set()} for m_id in MUNICIPIOS.keys()}
        new_count = 0
        progress_bar = st.progress(0)

        # 4. Process each PDF
        for i, pdf_file in enumerate(uploaded_pdfs):
            with pdfplumber.open(pdf_file) as pdf:
                text = "".join([p.extract_text() or "" for p in pdf.pages])
                tables = []
                for p in pdf.pages:
                    t = p.extract_table()
                    if t: tables.extend(t)

                dte_m = re.search(r'N[úu]mero\s*de\s*DTE:\s*(\d+)', text, re.IGNORECASE)
                dte_val = dte_m.group(1) if dte_m else pdf_file.name

                text_squished = squish_text(text)
                m_id, m_name = None, "N/A"
                
                # Check against our aggressively squished master list
                for alias, mun_id, official_name in search_list:
                    alias_squished = squish_text(alias)
                    if alias_squished in text_squished:
                        m_id = mun_id
                        m_name = official_name
                        break

                if m_id:
                    abar_sum, agri_sum = 0, 0
                    cultivados = ['tomate', 'pina', 'piña', 'banano', 'zanahoria', 'guisquil', 'güisquil', 'cebolla', 'aguacate', 
                                  'miltomate', 'brocoli', 'brócoli', 'melon', 'melón', 'ejote', 'maiz', 'maíz', 'jamaica', 
                                  'cebada', 'papaya', 'manzana', 'chile', 'apio', 'ajo', 'cilantro', 'tusa', 'sandia', 'sandía',
                                  'platano', 'plátano', 'naranja', 'limon', 'limón', 'lechuga', 'repollo', 'remolacha', 
                                  'rabano', 'rábano', 'pimiento', 'berenjena', 'calabaza', 'pepino']
                    abarrotes = ['pollo', 'tostada', 'huevo', 'pan', 'queso', 'carne', 'res', 'chowmein', 'chow mein', 
                                 'chaomein', 'chaumein', 'cahomein', 'crema', 'leche', 'mantequilla', 'aceite', 'arroz',
                                 'frijol', 'azucar', 'azúcar', 'sal', 'harina', 'pasta', 'fideos', 'atol', 'incaparina']
                    
                    # Find the Total column and Description column indices
                    total_col_idx = -1
                    desc_col_idx = -1
                    
                    for row_tbl in tables:
                        if not row_tbl: continue
                        for idx, cell in enumerate(row_tbl):
                            if not cell: continue
                            cell_norm = normalize_text(str(cell))
                            
                            # Find Total column (has "Total" and "(Q)")
                            if 'total' in cell_norm and 'descuento' not in cell_norm and '(q)' in cell_norm:
                                total_col_idx = idx
                            
                            # Find Description column
                            if 'descripcion' in cell_norm:
                                desc_col_idx = idx
                        
                        if total_col_idx != -1 and desc_col_idx != -1:
                            break
                    
                    # If we didn't find the description column, assume it's index 3
                    if desc_col_idx == -1:
                        desc_col_idx = 3

                    # Process each row in the tables
                    for row_tbl in tables:
                        if not row_tbl: continue
                        
                        # Build full row text for matching
                        row_text = " ".join([str(x) for x in row_tbl if x])
                        row_text_normalized = normalize_text(row_text)
                        
                        # FILTER 1: Skip rows with administrative keywords
                        skip_keywords = ['totales', 'superintendencia', 'datos del certificador', 
                                        'contribuyendo', 'sujeto a pagos', 'no genera derecho',
                                        'descripcion', 'cantidad', 'unitario', 'descuentos', 'impuestos']
                        if any(keyword in row_text_normalized for keyword in skip_keywords):
                            continue
                        
                        # FILTER 2: First cell should be a number (item number like 1, 2, 3...)
                        if row_tbl and row_tbl[0]:
                            first_cell = str(row_tbl[0]).strip()
                            # Check if first cell is a number (item rows start with 1, 2, 3, etc.)
                            if not first_cell.isdigit():
                                continue
                        else:
                            continue
                        
                        # Extract the value
                        val = extract_value_from_row(row_tbl, total_col_idx)
                        
                        # Skip rows with zero or invalid value
                        if val <= 0:
                            continue
                        
                        # Extract ONLY the description from the correct column
                        description = ""
                        if desc_col_idx < len(row_tbl) and row_tbl[desc_col_idx]:
                            description = str(row_tbl[desc_col_idx]).strip()
                        else:
                            # Fallback: try index 3
                            if len(row_tbl) > 3 and row_tbl[3]:
                                description = str(row_tbl[3]).strip()
                            else:
                                description = row_text
                        
                        # Use fuzzy matching to categorize (using full row text for matching)
                        category, matched_word = fuzzy_match_category(row_text, cultivados, abarrotes, threshold=80)
                        
                        if category == 'agricultura':
                            agri_sum += val
                        elif category == 'abarrotes':
                            abar_sum += val
                        elif category == 'unmatched':
                            # Add ONLY the description to unmatched items sheet
                            ws_unmatched.append([description, m_name, val, dte_val])
                    
                    nit_e_match = re.search(r'Emisor:\s*([0-9Kk\-]+)', text, re.I)
                    nit_r_match = re.search(r'Receptor:\s*([0-9Kk\-]+)', text, re.I)
                    name_e_match = re.search(r'(?:Factura(?:\s*Pequeño\s*Contribuyente)?)\s*\n+(.*?)\n+Nit\s*Emisor', text, re.IGNORECASE | re.DOTALL)
                    
                    nit_e = nit_e_match.group(1).strip() if nit_e_match else "N/A"
                    nit_r = nit_r_match.group(1).strip() if nit_r_match else "N/A"
                    school_name = extract_school_name(text)
                    raw_name = re.sub(r'\s+', ' ', name_e_match.group(1).strip() if name_e_match else "N/A")
                    name_e = re.split(r'(?i)n[úu]mero\s*de\s*autorizaci[óo]n', raw_name)[0]
                    name_e = re.split(r'(?i)\bserie\b', name_e)[0].strip()

                    batch_totals[m_id]['abar'] += abar_sum
                    batch_totals[m_id]['agri'] += agri_sum
                    if nit_e != "N/A": batch_totals[m_id]['emisores'].add(nit_e)
                    if nit_r != "N/A": batch_totals[m_id]['receptores'].add(nit_r)

                    total_rec = abar_sum + agri_sum
                    perc_abar = (abar_sum / total_rec) if total_rec > 0 else 0
                    alert_status = "⚠️ ALERTA: >30%" if perc_abar > 0.30 else "OK"

                    ws_det.append([name_e, nit_e, nit_r, school_name, dte_val, m_name, alert_status])
                    new_count += 1
                else:
                    st.warning(f"No se pudo identificar el municipio en la factura: {pdf_file.name}")

            progress_bar.progress((i + 1) / len(uploaded_pdfs))

        # 5. Write to Main Sheet securely
        for target_m_id, r_idx in row_map.items():
            data = batch_totals.get(target_m_id)
            if not data: continue

            if 'abar' in col_map and data['abar'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['abar'])
                target_cell.value = safe_float(target_cell.value) + data['abar']
            
            if 'agri' in col_map and data['agri'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['agri'])
                target_cell.value = safe_float(target_cell.value) + data['agri']

            if 'escuelas' in col_map and len(data['receptores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['escuelas'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['receptores'])
            
            if 'productores' in col_map and len(data['emisores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['productores'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['emisores'])

        # 6. Format "Extra Detalles" and "Items Sin Clasificar"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Format Extra Detalles
        for col in ws_det.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                cell.border = thin_border 
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_det.column_dimensions[col_letter].width = max_length + 2
        
        # Format Items Sin Clasificar
        for col in ws_unmatched.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                cell.border = thin_border 
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_unmatched.column_dimensions[col_letter].width = max_length + 2

        # 7. Final Export
        output = io.BytesIO()
        wb.save(output)
        
        # Count unmatched items (excluding header row)
        unmatched_count = ws_unmatched.max_row - 1 if ws_unmatched.max_row > 1 else 0
        
        success_msg = f"¡Proceso completado! {new_count} facturas procesadas y agregadas al Excel con éxito."
        if unmatched_count > 0:
            success_msg += f"""\n\n⚠️ {unmatched_count} items sin clasificar encontrados. Están en la tercera hoja del archivo de Excel, 'Items sin Clasificar', para revisión manual.
                            Los totales de esos productos no fueron agregados a la cantidad de la primera hoja"""
        
        st.success(success_msg)
        output.seek(0)
        st.download_button("Descargar Reporte Final", data=output.getvalue(), 
                           file_name="Reporte_MAGA_Actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Error crítico detectado: {e}")
